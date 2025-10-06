"""Translate COUNTIFS and SUMIFS formulas into Pandas snippets."""
from __future__ import annotations

import argparse
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import keyword
import math
import operator
import re as _re

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.utils.datetime import from_excel

from formula_parser import (
    BinaryOpNode,
    BooleanNode,
    ErrorNode,
    FormulaNode,
    FormulaParseError,
    FunctionNode,
    NameNode,
    NumberNode,
    ReferenceNode,
    StringNode,
    UnaryOpNode,
    parse_formula,
)


@dataclass
class FormulaCell:
    sheet: str
    address: str
    formula: str
    ast: Optional[FormulaNode] = None
    parse_error: Optional[str] = None


@dataclass
class FunctionCall:
    func_name: str
    args: List[str]
    raw_text: str
    offset: int = 0

    def __str__(self) -> str:  # pragma: no cover - debugging helper
        return f"{self.func_name}({', '.join(self.args)})"


@dataclass
class FilterExpression:
    sheet: str
    expression: str
    warnings: List[str] = field(default_factory=list)


@dataclass
class Translation:
    cell: FormulaCell
    expression: Optional[str]
    warnings: List[str] = field(default_factory=list)
    imports: set[str] = field(default_factory=set)


@dataclass
class TranslatorState:
    cell: FormulaCell
    context: WorkbookContext
    header_map: Dict[str, Dict[str, str]]
    warnings: List[str]
    imports: set[str]


@dataclass
class WorkbookContext:
    header_map: Dict[str, Dict[str, str]]
    cell_values: Dict[str, Dict[str, object]]


class TranslationError(Exception):
    """Raised when a formula cannot be translated into Python code."""
def load_workbook_context(path: str) -> tuple[WorkbookContext, List[FormulaCell]]:
    workbook = load_workbook(
        path,
        data_only=False,
        read_only=True,
        keep_links=False,
    )
    header_map: Dict[str, Dict[str, str]] = {}
    cell_values: Dict[str, Dict[str, object]] = {}
    formulas: List[FormulaCell] = []

    for ws in workbook.worksheets:
        sheet_name = ws.title
        headers: Dict[str, str] = {}
        values: Dict[str, object] = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for col_idx, cell in enumerate(row, start=1):
                column_letter = get_column_letter(col_idx)
                coordinate = f"{column_letter}{row_idx}".upper()
                value = cell.value
                values[coordinate] = value
                if row_idx == 1 and value is not None:
                    text = str(value).strip()
                    if text:
                        headers[column_letter] = text
                if getattr(cell, "data_type", None) == "f" and isinstance(value, str):
                    ast: Optional[FormulaNode] = None
                    parse_error: Optional[str] = None
                    try:
                        ast = parse_formula(value)
                    except FormulaParseError as exc:
                        parse_error = str(exc)
                    formulas.append(
                        FormulaCell(
                            sheet=sheet_name,
                            address=coordinate,
                            formula=value,
                            ast=ast,
                            parse_error=parse_error,
                        )
                    )
        header_map[sheet_name] = headers
        cell_values[sheet_name] = values

    workbook.close()
    return WorkbookContext(header_map, cell_values), formulas


def parse_string_literal(token: str) -> Optional[str]:
    if len(token) < 2 or not token.startswith('"') or not token.endswith('"'):
        return None
    inner = token[1:-1]
    return inner.replace('""', '"')




def df_reference(sheet: str) -> str:
    return f"dfs[{sheet!r}]"


def split_range_reference(reference: str, default_sheet: str) -> Optional[Tuple[str, str]]:
    token = reference.strip()
    if not token:
        return None
    sheet = default_sheet
    if "!" not in token:
        return sheet, token
    if token[0] == "'":
        idx = 1
        builder: List[str] = []
        while idx < len(token):
            ch = token[idx]
            if ch == "'":
                if idx + 1 < len(token) and token[idx + 1] == "'":
                    builder.append("'")
                    idx += 2
                    continue
                idx += 1
                break
            builder.append(ch)
            idx += 1
        sheet = "".join(builder)
        if idx < len(token) and token[idx] == "!":
            idx += 1
        ref_part = token[idx:]
        return sheet, ref_part
    sheet_part, ref_part = token.split("!", 1)
    return sheet_part, ref_part


COL_RE = re.compile(r"^([A-Z]+)(\d+)?$", re.IGNORECASE)
CELL_RE = re.compile(r"^([A-Z]+)(\d+)$", re.IGNORECASE)


def range_columns(range_token: str) -> Optional[List[str]]:
    token = range_token.replace("$", "")
    if "[" in token:
        return None
    if ":" in token:
        left, right = token.split(":", 1)
    else:
        left = right = token
    left_match = COL_RE.match(left)
    right_match = COL_RE.match(right)
    if not left_match or not right_match:
        return None
    start_idx = column_index_from_string(left_match.group(1))
    end_idx = column_index_from_string(right_match.group(1))
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return [get_column_letter(idx) for idx in range(start_idx, end_idx + 1)]


def range_to_column(
    range_ref: str,
    header_map: Dict[str, Dict[str, str]],
    default_sheet: str,
) -> Tuple[Optional[Tuple[str, str]], Optional[str]]:
    parsed = split_range_reference(range_ref, default_sheet)
    if not parsed:
        return None, f"Could not parse range reference: {range_ref!r}"
    sheet, ref = parsed
    columns = range_columns(ref)
    if not columns:
        return None, f"Unsupported range reference: {range_ref}"
    if len(columns) != 1:
        return None, f"Only single-column ranges are supported (saw {range_ref})"
    column_letter = columns[0]
    header = header_map.get(sheet, {}).get(column_letter)
    column_name = header if header is not None else column_letter
    return (sheet, column_name), None


def resolve_cell_reference(
    token: str,
    context: WorkbookContext,
    default_sheet: str,
) -> Tuple[bool, Optional[object]]:
    parsed = split_range_reference(token, default_sheet)
    if not parsed:
        return False, None
    sheet, ref = parsed
    if ":" in ref:
        return False, None
    cleaned = ref.replace("$", "")
    cell_match = CELL_RE.match(cleaned)
    if not cell_match:
        return False, None
    sheet_name = sheet
    sheet_values = context.cell_values.get(sheet_name)
    if sheet_values is None:
        return False, None
    key = cleaned.upper()
    value = sheet_values.get(key)
    if key not in sheet_values:
        return False, None
    return True, value


def coerce_literal(text: str) -> object:
    stripped = text.strip()
    if stripped.upper() in {"TRUE", "FALSE"}:
        return stripped.upper() == "TRUE"
    if stripped == "":
        return ""
    try:
        if "." in stripped:
            return float(stripped)
        return int(stripped)
    except ValueError:
        return stripped


def format_literal(value: object) -> str:
    if isinstance(value, str):
        return repr(value)
    if value is None:
        return "None"
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, (datetime, date)):
        return repr(value.isoformat())
    return repr(value)


def excel_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def normalize_cell_address(address: str) -> str:
    return address.replace("$", "").upper()


def reference_is_range(reference: str) -> bool:
    if ':' in reference:
        return True
    # Column-only references like "A:A" or "Sheet!A:A"
    cleaned = reference.replace('$', '')
    return not any(ch.isdigit() for ch in cleaned)


def resolve_scalar_reference(node: ReferenceNode, state: TranslatorState) -> object:
    if reference_is_range(node.reference):
        raise TranslationError(f"Range reference {node.original!r} cannot be used as a scalar")
    sheet = node.sheet or state.cell.sheet
    address = normalize_cell_address(node.reference)
    values = state.context.cell_values.get(sheet)
    if values is None or address not in values:
        raise TranslationError(f"Unable to resolve cell reference {node.original!r}")
    return values[address]


OPERATOR_MAP = [
    ("<>", "!="),
    (">=", ">="),
    ("<=", "<="),
    (">", ">"),
    ("<", "<"),
    ("=", "=="),
]

NUMBER_RE = re.compile(r"^-?\d+(?:\.\d+)?$")


def contains_wildcard(text: str) -> bool:
    return any(ch in text for ch in ("*", "?"))


def range_node_to_string(node: FormulaNode, state: TranslatorState) -> str:
    if isinstance(node, ReferenceNode):
        return node.original
    raise TranslationError("Expected a range reference")


def criteria_node_to_string(node: FormulaNode, state: TranslatorState) -> str:
    literal = try_evaluate_literal(node, state)
    if literal is not None:
        if isinstance(literal, str):
            return f'"{literal}"'
        return excel_text(literal)
    if isinstance(node, StringNode):
        return f'"{node.value}"'
    if isinstance(node, NumberNode):
        return node.text or str(node.value)
    if isinstance(node, BooleanNode):
        return "TRUE" if node.value else "FALSE"
    if isinstance(node, UnaryOpNode) and not node.postfix and node.operator in {'+', '-'}:
        operand = criteria_node_to_string(node.operand, state)
        return f"{node.operator}{operand}"
    if isinstance(node, BinaryOpNode) and node.operator == '&':
        left = criteria_node_to_string(node.left, state)
        right = criteria_node_to_string(node.right, state)
        return left + right
    raise TranslationError("Unsupported criteria expression")


def try_evaluate_literal(node: FormulaNode, state: TranslatorState) -> Optional[object]:
    try:
        return evaluate_literal(node, state)
    except TranslationError:
        return None


def evaluate_literal(node: FormulaNode, state: TranslatorState) -> object:
    if isinstance(node, NumberNode):
        if node.text and node.text.isdigit():
            return int(node.text)
        if node.value.is_integer():
            return int(node.value)
        return node.value
    if isinstance(node, StringNode):
        return node.value
    if isinstance(node, BooleanNode):
        return node.value
    if isinstance(node, ReferenceNode):
        return resolve_scalar_reference(node, state)
    if isinstance(node, UnaryOpNode):
        operand = evaluate_literal(node.operand, state)
        if node.postfix:
            if node.operator == '%':
                return operand / 100
            raise TranslationError(f"Unsupported postfix operator {node.operator!r}")
        if node.operator == '+':
            return +operand
        if node.operator == '-':
            return -operand
        raise TranslationError(f"Unsupported unary operator {node.operator!r}")
    if isinstance(node, BinaryOpNode):
        left = evaluate_literal(node.left, state)
        right = evaluate_literal(node.right, state)
        op = node.operator
        if op == '+':
            return left + right
        if op == '-':
            return left - right
        if op == '*':
            return left * right
        if op == '/':
            return left / right
        if op == '^':
            return math.pow(left, right)
        if op == '&':
            return f"{left}{right}"
        if op in {'=', '<>', '<', '>', '<=', '>='}:
            ops = {
                '=': operator.eq,
                '<>': operator.ne,
                '<': operator.lt,
                '>': operator.gt,
                '<=': operator.le,
                '>=': operator.ge,
            }
            return ops[op](left, right)
        raise TranslationError(f"Unsupported binary operator {op!r}")
    if isinstance(node, FunctionNode):
        name = node.name.upper()
        if name == 'EOMONTH':
            if len(node.args) != 2:
                raise TranslationError("EOMONTH requires two arguments")
            start = to_datetime_value(evaluate_literal(node.args[0], state))
            months = int(evaluate_literal(node.args[1], state))
            return eomonth(start, months)
        if name == 'YEAR':
            dt = to_datetime_value(evaluate_literal(node.args[0], state))
            return dt.year
        if name == 'MONTH':
            dt = to_datetime_value(evaluate_literal(node.args[0], state))
            return dt.month
        if name == 'IF':
            if len(node.args) < 2:
                raise TranslationError("IF requires at least two arguments")
            condition = evaluate_literal(node.args[0], state)
            if condition:
                return evaluate_literal(node.args[1], state)
            if len(node.args) >= 3:
                return evaluate_literal(node.args[2], state)
            return False
        raise TranslationError(f"Unsupported function {name} for literal evaluation")
    raise TranslationError(f"Unsupported node type {type(node).__name__}")


def to_datetime_value(value: object) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except TypeError as exc:
            raise TranslationError(f"Unable to convert Excel serial {value!r} to datetime") from exc
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError as exc:
            raise TranslationError(f"Unable to parse date string {value!r}") from exc
    raise TranslationError(f"Unsupported datetime value {value!r}")


def eomonth(start: datetime, months: int) -> datetime:
    month = start.month - 1 + months
    year = start.year + month // 12
    month = month % 12 + 1
    from calendar import monthrange

    last_day = monthrange(year, month)[1]
    return datetime(year, month, last_day, start.hour, start.minute, start.second)


def wildcard_to_regex(pattern: str) -> str:
    regex_parts: List[str] = ["^"]
    i = 0
    while i < len(pattern):
        ch = pattern[i]
        if ch == "*":
            regex_parts.append(".*")
        elif ch == "?":
            regex_parts.append(".")
        else:
            regex_parts.append(re.escape(ch))
        i += 1
    regex_parts.append("$")
    return "".join(regex_parts)


def build_filter_from_operator(lhs: str, op: str, payload: str) -> Tuple[Optional[str], List[str], Optional[str]]:
    warnings: List[str] = []
    if contains_wildcard(payload):
        if op not in {"==", "!="}:
            return None, warnings, f"Wildcards with operator {op} are not supported"
        regex = wildcard_to_regex(payload)
        expr = f"{lhs}.astype(str).str.match({regex!r})"
        if op == "!=":
            expr = f"~({expr})"
        return expr, warnings, None
    literal = coerce_literal(payload)
    if literal == "" and op == "!=":
        expr = f"~({lhs}.isna() | ({lhs} == \"\"))"
        return expr, warnings, None
    return f"{lhs} {op} {format_literal(literal)}", warnings, None


def split_concat_tokens(expr: str) -> List[str]:
    parts: List[str] = []
    token: List[str] = []
    depth = 0
    in_string = False
    i = 0
    while i < len(expr):
        ch = expr[i]
        if ch == '"':
            token.append(ch)
            if in_string:
                if i + 1 < len(expr) and expr[i + 1] == '"':
                    token.append('"')
                    i += 2
                    continue
                in_string = False
                i += 1
                continue
            in_string = True
            i += 1
            continue
        if not in_string:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(depth - 1, 0)
            elif ch == "&" and depth == 0:
                parts.append("".join(token).strip())
                token.clear()
                i += 1
                continue
        token.append(ch)
        i += 1
    if token:
        parts.append("".join(token).strip())
    return parts


def evaluate_concat_token(
    token: str,
    context: WorkbookContext,
    default_sheet: str,
) -> Tuple[Optional[str], List[str], Optional[str]]:
    warnings: List[str] = []
    stripped = token.strip()
    if stripped == "":
        return "", warnings, None
    literal = parse_string_literal(stripped)
    if literal is not None:
        return literal, warnings, None
    if NUMBER_RE.match(stripped):
        return str(coerce_literal(stripped)), warnings, None
    upper = stripped.upper()
    if upper in {"TRUE", "FALSE"}:
        return ("TRUE" if upper == "TRUE" else "FALSE"), warnings, None
    matched, value = resolve_cell_reference(stripped, context, default_sheet)
    if matched:
        return excel_text(value), warnings, None
    return None, warnings, f"Unsupported concatenation token: {token}"


def try_evaluate_concat(
    expr: str,
    context: WorkbookContext,
    default_sheet: str,
) -> Tuple[Optional[str], List[str], Optional[str]]:
    tokens = split_concat_tokens(expr)
    if len(tokens) <= 1:
        return None, [], None
    warnings: List[str] = []
    pieces: List[str] = []
    for token in tokens:
        value, token_warnings, error = evaluate_concat_token(token, context, default_sheet)
        warnings.extend(token_warnings)
        if error:
            return None, warnings, error
        if value is None:
            return None, warnings, f"Unable to evaluate concatenation token: {token}"
        pieces.append(value)
    return "".join(pieces), warnings, None


def build_filter_expression_from_string(
    lhs: str,
    value: str,
) -> Tuple[Optional[str], List[str], Optional[str]]:
    warnings: List[str] = []
    for token, py_op in OPERATOR_MAP:
        if value.startswith(token):
            remainder = value[len(token) :]
            expr, extra_warnings, error = build_filter_from_operator(lhs, py_op, remainder)
            warnings.extend(extra_warnings)
            return expr, warnings, error
    expr, extra_warnings, error = build_filter_from_operator(lhs, "==", value)
    warnings.extend(extra_warnings)
    return expr, warnings, error


def build_filter_expression(
    lhs: str,
    criteria: str,
    context: WorkbookContext,
    default_sheet: str,
) -> Tuple[Optional[str], List[str], Optional[str]]:
    crit = criteria.strip()
    warnings: List[str] = []
    if not crit:
        return None, warnings, "Empty criteria"
    concat_value, concat_warnings, concat_error = try_evaluate_concat(crit, context, default_sheet)
    warnings.extend(concat_warnings)
    if concat_error:
        return None, warnings, concat_error
    if concat_value is not None:
        expr, extra_warnings, error = build_filter_expression_from_string(lhs, concat_value)
        warnings.extend(extra_warnings)
        return expr, warnings, error
    literal = parse_string_literal(crit)
    if literal is not None:
        expr, extra_warnings, error = build_filter_expression_from_string(lhs, literal)
        warnings.extend(extra_warnings)
        return expr, warnings, error
    if NUMBER_RE.match(crit):
        return f"{lhs} == {coerce_literal(crit)}", warnings, None
    matched, value = resolve_cell_reference(crit, context, default_sheet)
    if matched:
        return f"{lhs} == {format_literal(value)}", warnings, None
    if crit.upper() in {"TRUE", "FALSE"}:
        literal = "True" if crit.upper() == "TRUE" else "False"
        return f"{lhs} == {literal}", warnings, None
    return None, warnings, f"Unsupported criteria: {criteria}"


def to_filter_expr(
    range_ref: str,
    criteria: str,
    context: WorkbookContext,
    header_map: Dict[str, Dict[str, str]],
    default_sheet: str,
) -> Tuple[Optional[FilterExpression], Optional[str]]:
    range_result, error = range_to_column(range_ref, header_map, default_sheet)
    if error:
        return None, error
    sheet, column_name = range_result
    lhs = f"{df_reference(sheet)}[{column_name!r}]"
    expression, warnings, crit_error = build_filter_expression(lhs, criteria, context, default_sheet)
    if crit_error:
        return None, crit_error
    return FilterExpression(sheet, expression, warnings), None


def combine_filters(filters: List[FilterExpression]) -> str:
    if not filters:
        return "slice(None)"
    joined = " & ".join(f"({flt.expression})" for flt in filters)
    return joined


def build_sumifs_translation(
    cell: FormulaCell,
    call: FunctionCall,
    context: WorkbookContext,
    header_map: Dict[str, Dict[str, str]],
) -> Tuple[Optional[str], List[str]]:
    warnings: List[str] = []
    if not call.args:
        return None, ["SUMIFS requires at least a sum range"]
    sum_range = call.args[0]
    sum_result, error = range_to_column(sum_range, header_map, cell.sheet)
    if error:
        return None, [error]
    sum_sheet, sum_column = sum_result
    criteria_args = call.args[1:]
    if len(criteria_args) % 2 != 0:
        return None, ["SUMIFS expects range/criteria pairs"]
    filters: List[FilterExpression] = []
    for range_ref, criteria in zip(criteria_args[::2], criteria_args[1::2]):
        filt, filt_error = to_filter_expr(range_ref, criteria, context, header_map, cell.sheet)
        if filt_error:
            return None, [filt_error]
        if filt is None:
            return None, ["Failed to build filter expression"]
        warnings.extend(filt.warnings)
        filters.append(filt)
    df_ref = df_reference(sum_sheet)
    if filters:
        mask_expr = combine_filters(filters)
        rows_expr = f"{df_ref}.loc[{mask_expr}]"
        expression = f"{rows_expr}[{sum_column!r}].sum()"
    else:
        expression = f"{df_ref}[{sum_column!r}].sum()"
    return expression, warnings


def build_countifs_translation(
    cell: FormulaCell,
    call: FunctionCall,
    context: WorkbookContext,
    header_map: Dict[str, Dict[str, str]],
) -> Tuple[Optional[str], List[str]]:
    warnings: List[str] = []
    if not call.args or len(call.args) % 2 != 0:
        return None, ["COUNTIFS expects an even number of arguments"]
    first_range = call.args[0]
    base_result, error = range_to_column(first_range, header_map, cell.sheet)
    if error:
        return None, [error]
    base_sheet, _ = base_result
    filters: List[FilterExpression] = []
    for range_ref, criteria in zip(call.args[::2], call.args[1::2]):
        filt, filt_error = to_filter_expr(range_ref, criteria, context, header_map, cell.sheet)
        if filt_error:
            return None, [filt_error]
        if filt is None:
            return None, ["Failed to build filter expression"]
        warnings.extend(filt.warnings)
        filters.append(filt)
    mask_expr = combine_filters(filters)
    df_ref = df_reference(base_sheet)
    expression = f"{df_ref}.loc[{mask_expr}].shape[0]"
    return expression, warnings


def translate_sumifs_node(node: FunctionNode, state: TranslatorState) -> str:
    args: List[str] = []
    for idx, arg in enumerate(node.args):
        if idx == 0 or idx % 2 == 1:
            args.append(range_node_to_string(arg, state))
        else:
            args.append(criteria_node_to_string(arg, state))
    call = FunctionCall(func_name='SUMIFS', args=args, raw_text='SUMIFS')
    expression, warnings = build_sumifs_translation(state.cell, call, state.context, state.header_map)
    state.warnings.extend(warnings)
    if expression is None:
        raise TranslationError("Unable to translate SUMIFS call")
    return expression


def translate_countifs_node(node: FunctionNode, state: TranslatorState) -> str:
    args: List[str] = []
    for idx, arg in enumerate(node.args):
        if idx % 2 == 0:
            args.append(range_node_to_string(arg, state))
        else:
            args.append(criteria_node_to_string(arg, state))
    call = FunctionCall(func_name='COUNTIFS', args=args, raw_text='COUNTIFS')
    expression, warnings = build_countifs_translation(state.cell, call, state.context, state.header_map)
    state.warnings.extend(warnings)
    if expression is None:
        raise TranslationError("Unable to translate COUNTIFS call")
    return expression


def translate_if_node(node: FunctionNode, state: TranslatorState) -> str:
    if len(node.args) < 2:
        raise TranslationError("IF requires at least two arguments")
    condition = translate_node(node.args[0], state)
    true_expr = translate_node(node.args[1], state)
    if len(node.args) >= 3:
        false_expr = translate_node(node.args[2], state)
    else:
        false_expr = "None"
    return f"({true_expr}) if ({condition}) else ({false_expr})"


def translate_eomonth_node(node: FunctionNode, state: TranslatorState) -> str:
    literal = try_evaluate_literal(node, state)
    if literal is not None:
        return format_literal(literal)
    raise TranslationError("EOMONTH with dynamic arguments is not supported yet")


def translate_year_node(node: FunctionNode, state: TranslatorState) -> str:
    literal = try_evaluate_literal(node, state)
    if literal is not None:
        return format_literal(literal)
    raise TranslationError("YEAR with dynamic arguments is not supported yet")


def translate_month_node(node: FunctionNode, state: TranslatorState) -> str:
    literal = try_evaluate_literal(node, state)
    if literal is not None:
        return format_literal(literal)
    raise TranslationError("MONTH with dynamic arguments is not supported yet")


def translate_index_node(node: FunctionNode, state: TranslatorState) -> str:
    if not node.args:
        raise TranslationError("INDEX expects at least one argument")
    range_str = range_node_to_string(node.args[0], state)
    range_result, error = range_to_column(range_str, state.header_map, state.cell.sheet)
    if error:
        raise TranslationError(error)
    target_sheet, target_column = range_result
    target_series = f"{df_reference(target_sheet)}[{target_column!r}]"

    if len(node.args) < 2:
        raise TranslationError("INDEX requires a row argument")
    row_arg = node.args[1]
    if isinstance(row_arg, FunctionNode) and row_arg.name.upper() == 'MATCH':
        lookup_expr = translate_match_node(row_arg, state)
        return f"{target_series}.loc[{lookup_expr}].iloc[0]"
    row_literal = try_evaluate_literal(row_arg, state)
    if row_literal is None:
        raise TranslationError("Unsupported INDEX row argument")
    row_index = int(row_literal) - 1
    return f"{target_series}.iloc[{row_index}]"


def translate_match_node(node: FunctionNode, state: TranslatorState) -> str:
    if len(node.args) < 2:
        raise TranslationError("MATCH requires at least two arguments")
    lookup_value = node.args[0]
    lookup_range_node = node.args[1]
    match_type = 0
    if len(node.args) >= 3:
        literal = try_evaluate_literal(node.args[2], state)
        if literal is None:
            raise TranslationError("Unsupported MATCH type")
        match_type = int(literal)
    if match_type != 0:
        raise TranslationError("Only exact MATCH (type 0) is supported")

    range_str = range_node_to_string(lookup_range_node, state)
    range_result, error = range_to_column(range_str, state.header_map, state.cell.sheet)
    if error:
        raise TranslationError(error)
    sheet, column = range_result
    series = f"{df_reference(sheet)}[{column!r}]"

    literal = try_evaluate_literal(lookup_value, state)
    if literal is not None:
        value_expr = format_literal(literal)
    else:
        value_expr = translate_node(lookup_value, state)
    return f"({series} == {value_expr})"


def translate_node(node: FormulaNode, state: TranslatorState) -> str:
    if isinstance(node, NumberNode):
        return node.text or str(node.value)
    if isinstance(node, StringNode):
        return repr(node.value)
    if isinstance(node, BooleanNode):
        return "True" if node.value else "False"
    if isinstance(node, ErrorNode):
        raise TranslationError(f"Error literal {node.value!r} is not supported")
    if isinstance(node, ReferenceNode):
        value = resolve_scalar_reference(node, state)
        return format_literal(value)
    if isinstance(node, NameNode):
        raise TranslationError(f"Named ranges are not supported ({node.name})")
    if isinstance(node, UnaryOpNode):
        operand = translate_node(node.operand, state)
        if node.postfix:
            if node.operator == '%':
                return f"({operand}) / 100"
            raise TranslationError(f"Unsupported postfix operator {node.operator!r}")
        if node.operator in {'+', '-'}:
            return f"({node.operator}{operand})"
        raise TranslationError(f"Unsupported unary operator {node.operator!r}")
    if isinstance(node, BinaryOpNode):
        left = translate_node(node.left, state)
        right = translate_node(node.right, state)
        op = node.operator
        if op == '^':
            return f"({left}) ** ({right})"
        if op == '&':
            return f"({left}) + ({right})"
        if op == '=':
            op = '=='
        return f"({left}) {op} ({right})"
    if isinstance(node, FunctionNode):
        name = node.name.upper()
        if name == 'SUMIFS':
            return translate_sumifs_node(node, state)
        if name == 'COUNTIFS':
            return translate_countifs_node(node, state)
        if name == 'IF':
            return translate_if_node(node, state)
        if name == 'EOMONTH':
            return translate_eomonth_node(node, state)
        if name == 'YEAR':
            return translate_year_node(node, state)
        if name == 'MONTH':
            return translate_month_node(node, state)
        if name == 'INDEX':
            return translate_index_node(node, state)
        if name == 'MATCH':
            return translate_match_node(node, state)
        raise TranslationError(f"Unsupported function {name}")
    raise TranslationError(f"Unsupported node type {type(node).__name__}")


def translate_call(
    cell: FormulaCell,
    call: FunctionCall,
    context: WorkbookContext,
    header_map: Dict[str, Dict[str, str]],
) -> Translation:
    if call.func_name == "SUMIFS":
        expression, warnings = build_sumifs_translation(cell, call, context, header_map)
    else:
        expression, warnings = build_countifs_translation(cell, call, context, header_map)
    return Translation(cell, call, expression, warnings)


def translate_workbook(
    context: WorkbookContext,
    formula_cells: Iterable[FormulaCell],
) -> Iterator[Translation]:
    header_map = context.header_map
    for cell in formula_cells:
        state = TranslatorState(cell=cell, context=context, header_map=header_map, warnings=[], imports=set())
        if cell.parse_error:
            state.warnings.append(f"Parse error: {cell.parse_error}")
            yield Translation(cell=cell, expression=None, warnings=state.warnings, imports=state.imports)
            continue
        if cell.ast is None:
            state.warnings.append("No AST available")
            yield Translation(cell=cell, expression=None, warnings=state.warnings, imports=state.imports)
            continue
        try:
            expression = translate_node(cell.ast, state)
        except TranslationError as exc:
            state.warnings.append(str(exc))
            expression = None
        yield Translation(cell=cell, expression=expression, warnings=state.warnings, imports=state.imports)


_identifier_re = _re.compile(r"[^0-9a-zA-Z_]+")


def _sanitize_identifier(sheet: str, address: str, existing: set[str]) -> str:
    base = f"{sheet}_{address}".lower()
    base = base.replace("!", "_")
    base = _identifier_re.sub("_", base)
    if not base:
        base = "formula"
    if base[0].isdigit():
        base = f"cell_{base}"
    if keyword.iskeyword(base):
        base = f"formula_{base}"
    candidate = base
    counter = 1
    while candidate in existing:
        counter += 1
        candidate = f"{base}_{counter}"
    existing.add(candidate)
    return candidate


def write_formula_module(
    translations: List[Translation],
    output_path: Path,
    workbook_name: str,
) -> None:
    existing: set[str] = set()
    lines: List[str] = []
    lines.append(f'"""Auto-generated pandas formulas from {workbook_name}."""')
    lines.append("")
    imports: set[str] = set()
    for translation in translations:
        imports.update(translation.imports)
    if imports:
        for statement in sorted(imports):
            lines.append(statement)
        lines.append("")
    lines.append("# Expect a dict-like object `dfs` mapping sheet names to pandas DataFrames.")
    lines.append("# Example: dfs = {'Sales': sales_df, 'Summary': summary_df}")
    lines.append("# Evaluate these expressions after populating `dfs` to obtain the metrics.")
    lines.append("")

    for translation in translations:
        cell = translation.cell
        identifier = _sanitize_identifier(cell.sheet, cell.address, existing)
        lines.append(f"# {cell.sheet}!{cell.address}: {cell.formula}")
        if cell.parse_error:
            lines.append(f"# WARNING: parse error -> {cell.parse_error}")
        for warning in translation.warnings:
            lines.append(f"# NOTE: {warning}")
        if translation.expression:
            lines.append(f"{identifier} = {translation.expression}")
        else:
            lines.append("# Unable to translate this formula into pandas code.")
        lines.append("")

    output_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to the Excel workbook to inspect")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("generated_formulas.py"),
        help="Destination .py file for the generated pandas code",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress information while processing the workbook.",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    context, formulas = load_workbook_context(args.workbook)
    if args.verbose:
        if not formulas:
            print("No COUNTIFS or SUMIFS formulas found in the workbook.")
        else:
            counts = Counter(cell.sheet for cell in formulas)
            total = sum(counts.values())
            print(f"Discovered {total} target formula cell(s) across {len(counts)} sheet(s):")
            for sheet, count in counts.items():
                print(f"  {sheet}: {count}")

    translations = list(translate_workbook(context, formulas))
    if not translations:
        print("No COUNTIFS or SUMIFS formulas found.")
        return 0

    output_path = args.output if args.output.is_absolute() else Path.cwd() / args.output
    write_formula_module(translations, output_path, Path(args.workbook).name)
    if args.verbose:
        print("Generated expressions:")
        for translation in translations:
            status = "ok" if translation.expression else "skipped"
            print(f"  {translation.cell.sheet}!{translation.cell.address} -> {status}")
    print(f"Generated pandas formulas written to {output_path}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
